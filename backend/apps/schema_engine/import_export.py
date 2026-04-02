from __future__ import annotations

import csv
import io
from typing import AsyncGenerator

from fastapi import UploadFile
from fastapi.responses import StreamingResponse

from apps.schema_engine.models import FieldType
from apps.schema_engine.repository import get_model_by_slug
from apps.schema_engine.service import create_record, list_records


# ── Field type helpers ───────────────────────────────────────────────────────

# Field types that cannot be represented in a flat CSV
_SKIP_TYPES = {
    FieldType.FORMULA, FieldType.ROLLUP, FieldType.CHILD_TABLE,
    FieldType.SECTION_BREAK, FieldType.COLUMN_BREAK, FieldType.PAGE_BREAK,
    FieldType.SIGNATURE, FieldType.GEOLOCATION, FieldType.JSON,
}


def _exportable_fields(model) -> list:
    """Return ordered list of exportable field definitions."""
    return [
        f for f in sorted(model.fields, key=lambda x: x.order)
        if f.type not in _SKIP_TYPES
    ]


def _field_names(model) -> list[str]:
    """Return ordered list of exportable field names."""
    return [f.name for f in _exportable_fields(model)]


def _format_value(val, field) -> str:
    """Format a record value for CSV/XLSX export based on field type."""
    if val is None:
        return ""
    if field.type in (FieldType.MULTISELECT, FieldType.TABLE_MULTISELECT):
        if isinstance(val, list):
            return ", ".join(str(v) for v in val)
        return str(val)
    if field.type == FieldType.BOOLEAN:
        return "true" if val else "false"
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    return str(val)


def _sample_value(field) -> str:
    """Generate a sample value for a field type (used in import templates)."""
    t = field.type
    if t == FieldType.TEXT:
        return "Sample Text"
    if t == FieldType.NUMBER:
        return "100"
    if t == FieldType.CURRENCY:
        return "99.99"
    if t == FieldType.DATE:
        return "2024-01-15"
    if t == FieldType.DATETIME:
        return "2024-01-15T10:30:00"
    if t == FieldType.TIME:
        return "10:30:00"
    if t == FieldType.DURATION:
        return "01:30:00"
    if t == FieldType.EMAIL:
        return "user@example.com"
    if t == FieldType.URL:
        return "https://example.com"
    if t == FieldType.PHONE:
        return "+1234567890"
    if t == FieldType.BOOLEAN:
        return "true"
    if t == FieldType.SELECT:
        options = field.config.get("options", [])
        if options:
            opt = options[0]
            return opt.get("value", opt) if isinstance(opt, dict) else str(opt)
        return "Option1"
    if t == FieldType.MULTISELECT:
        options = field.config.get("options", [])
        if options:
            vals = []
            for opt in options[:2]:
                vals.append(opt.get("value", opt) if isinstance(opt, dict) else str(opt))
            return ", ".join(vals)
        return "Option1, Option2"
    if t == FieldType.RELATION:
        return "<related_record_id>"
    if t == FieldType.TABLE_MULTISELECT:
        return "<record_id_1>, <record_id_2>"
    if t == FieldType.USER_REF:
        return "<user_id>"
    if t == FieldType.RATING:
        return "4"
    if t == FieldType.COLOR:
        return "#3B82F6"
    if t == FieldType.BARCODE:
        return "ABC-123-XYZ"
    if t == FieldType.RICH_TEXT:
        return "Sample rich text content"
    if t == FieldType.HTML:
        return "<p>Sample HTML</p>"
    if t == FieldType.ICON:
        return "star"
    if t == FieldType.DYNAMIC_LINK:
        return "ModelSlug/record_id"
    if t == FieldType.FILE or t == FieldType.ATTACH_IMAGE:
        return ""
    return "sample"


def _field_type_hint(field) -> str:
    """Return a human-readable type hint for column mapping UI."""
    t = field.type
    hint = t.value
    if t == FieldType.SELECT:
        options = field.config.get("options", [])
        vals = [opt.get("value", opt) if isinstance(opt, dict) else str(opt) for opt in options[:5]]
        if vals:
            hint += f" ({', '.join(vals)})"
    elif t == FieldType.DATE:
        hint = "date (YYYY-MM-DD)"
    elif t == FieldType.DATETIME:
        hint = "datetime (YYYY-MM-DDTHH:MM:SS)"
    elif t == FieldType.BOOLEAN:
        hint = "boolean (true/false)"
    elif t == FieldType.MULTISELECT:
        options = field.config.get("options", [])
        vals = [opt.get("value", opt) if isinstance(opt, dict) else str(opt) for opt in options[:5]]
        if vals:
            hint += f" ({', '.join(vals)})"
    if field.is_required:
        hint += " *required"
    return hint


# ── Import Template ──────────────────────────────────────────────────────────

async def generate_import_template(
    model_slug: str,
    app_id: str,
    format: str = "csv",
) -> StreamingResponse:
    """Generate a downloadable import template with headers and a sample row."""
    model = await get_model_by_slug(model_slug, app_id)
    if model is None:
        from core.exceptions import NotFoundError
        raise NotFoundError("Model", model_slug)

    fields = _exportable_fields(model)
    headers = [f.label or f.name for f in fields]
    sample_row = [_sample_value(f) for f in fields]

    if format == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{model_slug}_template"[:31]

        # Header row
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Sample row
        ws.append(sample_row)
        sample_font = Font(italic=True, color="888888")
        for cell in ws[2]:
            cell.font = sample_font

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{model_slug}_template.xlsx"'},
        )

    # Default: CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(sample_row)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{model_slug}_template.csv"'},
    )


# ── Export ────────────────────────────────────────────────────────────────────

def _record_to_row(record: dict, fields: list) -> list:
    data = record.get("data", record)
    return [_format_value(data.get(f.name), f) for f in fields]


async def export_csv(
    model_slug: str,
    app_id: str,
    tenant_id: str,
) -> StreamingResponse:
    model = await get_model_by_slug(model_slug, app_id)
    if model is None:
        from core.exceptions import NotFoundError
        raise NotFoundError("Model", model_slug)

    fields = _exportable_fields(model)
    field_labels = [f.label or f.name for f in fields]

    async def generate() -> AsyncGenerator[str, None]:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["_id"] + field_labels)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        page = 1
        while True:
            result = await list_records(
                model_slug=model_slug, app_id=app_id, tenant_id=tenant_id,
                page=page, page_size=200, sort_field="created_at", sort_dir=-1,
            )
            for rec in result.items:
                writer.writerow([str(rec.get("id", rec.get("_id", "")))] + _record_to_row(rec, fields))
                yield output.getvalue()
                output.seek(0)
                output.truncate(0)
            if not result.has_more:
                break
            page += 1

    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{model_slug}.csv"'},
    )


async def export_xlsx(
    model_slug: str,
    app_id: str,
    tenant_id: str,
) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    model = await get_model_by_slug(model_slug, app_id)
    if model is None:
        from core.exceptions import NotFoundError
        raise NotFoundError("Model", model_slug)

    fields = _exportable_fields(model)
    field_labels = [f.label or f.name for f in fields]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_slug[:31]

    # Header row
    headers = ["_id"] + field_labels
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    page = 1
    while True:
        result = await list_records(
            model_slug=model_slug, app_id=app_id, tenant_id=tenant_id,
            page=page, page_size=500, sort_field="created_at", sort_dir=-1,
        )
        for rec in result.items:
            ws.append([str(rec.get("id", rec.get("_id", "")))] + _record_to_row(rec, fields))
        if not result.has_more:
            break
        page += 1

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{model_slug}.xlsx"'},
    )


# ── Relation Lookup Builder (shared by CSV and Google Sheets import) ──────────

async def _build_relation_lookups(
    mapping: dict[str, str],
    model_fields: dict,
    app_id: str,
    tenant_id: str,
) -> dict[str, dict[str, str]]:
    """Pre-build lookup caches for relation fields.

    For each relation field in the mapping, fetches records from the related
    model and builds {display_value_lowercase: record_id}.
    This lets users import "SI" instead of raw ObjectIds.
    """
    relation_lookups: dict[str, dict[str, str]] = {}

    for field_name in mapping.values():
        field = model_fields.get(field_name)
        if not field or field.type != FieldType.RELATION:
            continue
        related_slug = field.config.get("related_model_slug")
        target_app_id = field.config.get("target_app_id") or app_id
        if not related_slug:
            continue
        try:
            related_model = await get_model_by_slug(related_slug, target_app_id)
            if not related_model:
                continue
            # Determine display field: config.display_field → name/title → first text field
            display_field = field.config.get("display_field")
            if not display_field:
                for rf in related_model.fields:
                    if rf.type == FieldType.TEXT and rf.name in ("name", "title"):
                        display_field = rf.name
                        break
                if not display_field:
                    for rf in related_model.fields:
                        if rf.type == FieldType.TEXT:
                            display_field = rf.name
                            break
            if not display_field:
                display_field = "name"

            # Fetch related records (up to 5000)
            from core.database import get_tenant_db
            db = get_tenant_db(tenant_id)
            col = db[f"data__{target_app_id}__{related_slug}"]
            cursor = col.find({}, {"_id": 1, display_field: 1}).limit(5000)
            lookup = {}
            async for doc in cursor:
                val = doc.get(display_field, "")
                if val:
                    lookup[str(val).strip().lower()] = str(doc["_id"])
            relation_lookups[field_name] = lookup
        except Exception:
            pass  # If lookup fails, user must provide IDs directly

    return relation_lookups


# ── Import ────────────────────────────────────────────────────────────────────

async def import_csv(
    model_slug: str,
    app_id: str,
    tenant_id: str,
    file: UploadFile,
    user_id: str,
    column_map: dict[str, str] | None = None,
) -> dict:
    """
    Parse uploaded CSV and create records.
    column_map: {csv_column_name -> model_field_name}  (None = auto-match by name)
    Returns {"created": N, "failed": N, "errors": [...]}
    """
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")  # handle BOM
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        return {"created": 0, "failed": 0, "errors": [{"row": 0, "field": "", "message": "Empty or invalid CSV file"}]}

    model = await get_model_by_slug(model_slug, app_id)
    if model is None:
        from core.exceptions import NotFoundError
        raise NotFoundError("Model", model_slug)

    model_fields = {f.name: f for f in model.fields}
    # Also build label->name map for matching by label
    label_to_name = {}
    for f in model.fields:
        if f.label:
            label_to_name[f.label.strip().lower()] = f.name

    # Build mapping: csv column -> field name
    if column_map:
        mapping = {}
        for csv_col, field_name in column_map.items():
            if field_name in model_fields:
                mapping[csv_col] = field_name
            elif field_name.strip().lower() in label_to_name:
                mapping[csv_col] = label_to_name[field_name.strip().lower()]
    else:
        # Auto-map: match by name or label (case-insensitive), skip _id
        mapping = {}
        for csv_col in (reader.fieldnames or []):
            if csv_col == "_id":
                continue
            normalized = csv_col.strip().lower().replace(" ", "_")
            # Try name match
            matched = False
            for fname in model_fields:
                if fname.lower() == normalized:
                    mapping[csv_col] = fname
                    matched = True
                    break
            # Try label match
            if not matched:
                label_key = csv_col.strip().lower()
                if label_key in label_to_name:
                    mapping[csv_col] = label_to_name[label_key]

    # Pre-build relation lookup caches for display-value → ID resolution
    relation_lookups = await _build_relation_lookups(mapping, model_fields, app_id, tenant_id)

    created = 0
    failed = 0
    errors: list[dict] = []

    for i, row in enumerate(reader, start=2):  # row 1 = header
        data: dict = {}
        row_errors = []
        for csv_col, field_name in mapping.items():
            raw = row.get(csv_col, "").strip()
            if not raw:
                continue
            field = model_fields[field_name]
            # Type coercion
            try:
                if field.type in (FieldType.NUMBER, FieldType.CURRENCY):
                    data[field_name] = float(raw)
                elif field.type == FieldType.BOOLEAN:
                    data[field_name] = raw.lower() in ("true", "1", "yes")
                elif field.type == FieldType.MULTISELECT:
                    data[field_name] = [v.strip() for v in raw.split(",") if v.strip()]
                elif field.type == FieldType.TABLE_MULTISELECT:
                    data[field_name] = [v.strip() for v in raw.split(",") if v.strip()]
                elif field.type == FieldType.RATING:
                    data[field_name] = int(float(raw))
                elif field.type == FieldType.RELATION:
                    # Resolve display value to record ID
                    lookup = relation_lookups.get(field_name, {})
                    resolved = lookup.get(raw.strip().lower())
                    if resolved:
                        data[field_name] = resolved
                    elif len(raw) == 24 and all(c in "0123456789abcdef" for c in raw):
                        # Looks like an ObjectId — use as-is
                        data[field_name] = raw
                    else:
                        row_errors.append({
                            "row": i, "field": field_name,
                            "message": f"Cannot resolve '{raw}' — no matching record found in related model"
                        })
                        continue
                elif field.type == FieldType.SELECT:
                    # Validate against options — tolerant matching
                    options = field.config.get("options", [])
                    valid_values = []
                    for opt in options:
                        if isinstance(opt, dict):
                            valid_values.append(opt.get("value", ""))
                        else:
                            valid_values.append(str(opt))
                    if valid_values:
                        # 1. Exact match
                        if raw in valid_values:
                            data[field_name] = raw
                        else:
                            # 2. Case-insensitive + whitespace-stripped match
                            raw_norm = raw.strip().lower()
                            matched_opt = None
                            for v in valid_values:
                                if v.strip().lower() == raw_norm:
                                    matched_opt = v
                                    break
                            if matched_opt:
                                data[field_name] = matched_opt
                            else:
                                # 3. Partial/contains match (e.g., "NA" matches "N/A")
                                raw_alpha = "".join(c for c in raw_norm if c.isalnum())
                                for v in valid_values:
                                    v_alpha = "".join(c for c in v.strip().lower() if c.isalnum())
                                    if raw_alpha == v_alpha:
                                        matched_opt = v
                                        break
                                if matched_opt:
                                    data[field_name] = matched_opt
                                else:
                                    row_errors.append({
                                        "row": i, "field": field_name,
                                        "message": f"Invalid option '{raw}'. Valid: {', '.join(valid_values[:5])}"
                                    })
                                    continue
                    else:
                        data[field_name] = raw
                else:
                    data[field_name] = raw
            except (ValueError, TypeError) as e:
                row_errors.append({
                    "row": i, "field": field_name,
                    "message": f"Type error for '{field_name}': {str(e)}"
                })

        if row_errors:
            errors.extend(row_errors)
            failed += 1
            continue

        if not data:
            continue

        # Check required fields
        missing = []
        for fname, fdef in model_fields.items():
            if fdef.is_required and fname not in data and fdef.type not in _SKIP_TYPES:
                missing.append(fdef.label or fname)
        if missing:
            errors.append({
                "row": i, "field": ", ".join(missing),
                "message": f"Required field(s) missing: {', '.join(missing)}"
            })
            failed += 1
            continue

        try:
            await create_record(model_slug, app_id, tenant_id, data, user_id=user_id)
            created += 1
        except Exception as exc:
            failed += 1
            errors.append({"row": i, "field": "", "message": str(exc)})
            if len(errors) >= 50:  # cap error list
                errors.append({"row": 0, "field": "", "message": "... more errors truncated"})
                break

    return {"created": created, "failed": failed, "errors": errors}


# ── Google Sheets Import ──────────────────────────────────────────────────────

import re
import httpx


def _extract_sheet_id(url: str) -> tuple[str, str]:
    """Extract the spreadsheet ID and gid from a Google Sheets URL.

    Supports formats:
      - https://docs.google.com/spreadsheets/d/SHEET_ID/edit#gid=0
      - https://docs.google.com/spreadsheets/d/SHEET_ID/edit?gid=0
      - https://docs.google.com/spreadsheets/d/SHEET_ID/
    Returns (sheet_id, gid).
    """
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if not match:
        raise ValueError("Invalid Google Sheets URL. Please provide a valid sharing link.")
    sheet_id = match.group(1)

    gid_match = re.search(r'[#?&]gid=(\d+)', url)
    gid = gid_match.group(1) if gid_match else "0"

    return sheet_id, gid


async def fetch_google_sheet_csv(url: str) -> str:
    """Fetch CSV content from a public/shared Google Sheets URL."""
    sheet_id, gid = _extract_sheet_id(url)
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

    async with httpx.AsyncClient(follow_redirects=True, timeout=30.0) as client:
        resp = await client.get(export_url)
        if resp.status_code == 404:
            raise ValueError("Spreadsheet not found. Make sure the sheet exists and sharing is enabled.")
        if resp.status_code == 403:
            raise ValueError("Access denied. Make sure the sheet is shared as 'Anyone with the link can view'.")
        if resp.status_code != 200:
            raise ValueError(f"Failed to fetch sheet (HTTP {resp.status_code}). Check the URL and sharing settings.")
        return resp.text


async def import_google_sheet(
    model_slug: str,
    app_id: str,
    tenant_id: str,
    sheet_url: str,
    user_id: str,
    column_map: dict[str, str] | None = None,
) -> dict:
    """
    Import records from a Google Sheets URL.
    The sheet must be shared as 'Anyone with the link can view'.
    Uses the same import logic as CSV import.
    """
    csv_text = await fetch_google_sheet_csv(sheet_url)

    # Parse headers for the frontend mapping step if no column_map provided
    reader = csv.DictReader(io.StringIO(csv_text))
    if reader.fieldnames is None:
        return {"created": 0, "failed": 0, "errors": [{"row": 0, "field": "", "message": "Empty spreadsheet"}]}

    if column_map is None:
        # Return headers for the frontend to show the mapping step
        cleaned = [h.strip() for h in reader.fieldnames if h.strip() and h.strip() != '_id']
        return {"step": "map", "headers": cleaned, "row_count": sum(1 for _ in reader)}

    # Full import with column mapping
    model = await get_model_by_slug(model_slug, app_id)
    if model is None:
        from core.exceptions import NotFoundError
        raise NotFoundError("Model", model_slug)

    model_fields = {f.name: f for f in model.fields}
    label_to_name = {}
    for f in model.fields:
        if f.label:
            label_to_name[f.label.strip().lower()] = f.name

    # Build mapping
    mapping = {}
    for csv_col, field_name in column_map.items():
        if field_name in model_fields:
            mapping[csv_col] = field_name
        elif field_name.strip().lower() in label_to_name:
            mapping[csv_col] = label_to_name[field_name.strip().lower()]

    # Build relation lookups (same as CSV import)
    relation_lookups = await _build_relation_lookups(mapping, model_fields, app_id, tenant_id)

    created = 0
    failed = 0
    errors: list[dict] = []

    for i, row in enumerate(reader, start=2):
        data: dict = {}
        row_errors = []
        for csv_col, field_name in mapping.items():
            raw = row.get(csv_col, "").strip()
            if not raw:
                continue
            field = model_fields[field_name]
            try:
                if field.type in (FieldType.NUMBER, FieldType.CURRENCY):
                    data[field_name] = float(raw)
                elif field.type == FieldType.BOOLEAN:
                    data[field_name] = raw.lower() in ("true", "1", "yes")
                elif field.type == FieldType.MULTISELECT:
                    data[field_name] = [v.strip() for v in raw.split(",") if v.strip()]
                elif field.type == FieldType.RATING:
                    data[field_name] = int(float(raw))
                elif field.type == FieldType.RELATION:
                    lookup = relation_lookups.get(field_name, {})
                    resolved = lookup.get(raw.strip().lower())
                    if resolved:
                        data[field_name] = resolved
                    elif len(raw) == 24 and all(c in "0123456789abcdef" for c in raw):
                        data[field_name] = raw
                    else:
                        row_errors.append({
                            "row": i, "field": field_name,
                            "message": f"Cannot resolve '{raw}' — no matching record found in related model"
                        })
                        continue
                else:
                    data[field_name] = raw
            except (ValueError, TypeError) as e:
                row_errors.append({"row": i, "field": field_name, "message": str(e)})

        if row_errors:
            errors.extend(row_errors)
            failed += 1
            continue
        if not data:
            continue

        try:
            await create_record(model_slug, app_id, tenant_id, data, user_id=user_id)
            created += 1
        except Exception as exc:
            failed += 1
            errors.append({"row": i, "field": "", "message": str(exc)})
            if len(errors) >= 50:
                errors.append({"row": 0, "field": "", "message": "... more errors truncated"})
                break

    return {"created": created, "failed": failed, "errors": errors}


async def get_field_type_hints(model_slug: str, app_id: str) -> list[dict]:
    """Return field type hints for the import column mapping UI."""
    model = await get_model_by_slug(model_slug, app_id)
    if model is None:
        return []
    fields = _exportable_fields(model)
    return [
        {
            "name": f.name,
            "label": f.label or f.name,
            "type": f.type.value,
            "hint": _field_type_hint(f),
            "is_required": f.is_required,
        }
        for f in fields
    ]
